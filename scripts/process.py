import pandas as pd
from IPython.display import clear_output
from pandas.errors import ParserError
import os
from openpyxl import load_workbook
from tqdm import tqdm
from datetime import datetime
import numpy as np
import re



# Function to clean the dataframe
def data_import(file_name, data_dir):
    
    """
    Import a DataFrame.
    """
    
    # Determine the file path and extension
    for ext in ['.csv', '.xlsx', '.xls']:
        file_path = os.path.join(data_dir, f'{file_name}{ext}')
        if os.path.exists(file_path):
            break
    else:
        raise FileNotFoundError(f"No file found for {file_name} with supported extensions (.csv, .xlsx, .xls)")

    # Load dataset based on file type
    
    print(f"Importing data: {file_name}{ext} ---")
    
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path, low_memory=False)
    elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        while True:
            sheet_name = input("Please enter the sheet name: ")
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                break
            except ValueError:
                print("Sheet name does not exist. Please try again.")
    else:
        raise ValueError("Unsupported file format")
    
    print(f"Imported {file_name}{ext} successfully")

    dict_path = create_dictionary(df, file_name, data_dir)

    # Return the final cleaned and transformed dataframe
    return df, dict_path



def create_dictionary(df, file_name, data_dir):
    
    """
    Creates a data dictionary from a DataFrame.
    """
    
    data_dict = {
        'Include': [],
        'Index': [],
        'Source Name': [],
        'Column Name': [],
        'Column Label': [],
        'Data Type': [],
        'Required': [],
        'Unique': [],
        'Range': [],
        'Min': [],
        'Max': [],
        'Regex Pattern': [],
        'Default Value': [],
        'Null Value': [],
        'Format': [],
        'Examples': []
    }
    
    for index, col in tqdm(enumerate(df.columns), total=len(df.columns), desc="Creating dictionary"):
        data_dict['Index'].append(index+1)
        data_dict['Source Name'].append(col)
        data_dict['Data Type'].append(str(df[col].dtype))
        data_dict['Required'].append(not df[col].isnull().any())
        data_dict['Unique'].append(df[col].nunique() == len(df[col]))
        data_dict['Examples'].append(df[col].dropna().unique()[:5])
        
        # Determine the range for numerical and categorical data
        if pd.api.types.is_numeric_dtype(df[col]):
            data_dict['Range'].append((df[col].min(), df[col].max()))
            data_dict['Min'].append(df[col].min())
            data_dict['Max'].append(df[col].max())
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            data_dict['Range'].append((df[col].min(), df[col].max()))
            data_dict['Min'].append(df[col].min())
            data_dict['Max'].append(df[col].max())
        else:
            unique_vals = sorted(map(str, df[col].dropna().unique()))
            if len(unique_vals) <= 10:  # Arbitrarily choosing 10 as a cutoff for display
                data_dict['Range'].append(f"Categorical with {len(unique_vals)} unique values: \n {unique_vals}")
                data_dict['Min'].append(None)
                data_dict['Max'].append(None)
            else:
                data_dict['Range'].append(f"Categorical with {len(unique_vals)} unique values: \n {unique_vals[:5]} etc.")
                data_dict['Min'].append(None)
                data_dict['Max'].append(None)
        
        # Append None for other properties
        data_dict['Include'].append(None)
        data_dict['Column Name'].append(None)
        data_dict['Column Label'].append(None)
        data_dict['Regex Pattern'].append(None)
        data_dict['Default Value'].append(None)
        data_dict['Null Value'].append(None)
        data_dict['Format'].append(None)

    # Convert the dictionary to a DataFrame
    dict = pd.DataFrame(data_dict)

    # Define the path for the Excel file
    excel_path = f"{data_dir}/data_dict-{file_name}.xlsx"

    sheet_name = f"Dict-{file_name}_update"
    
    # Open the existing workbook using openpyxl
    if os.path.exists(excel_path):
        book = load_workbook(excel_path)
        # Remove the existing sheet if it exists
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        # Save the changes
        book.save(excel_path)
        book.close()
    
    # Check if the file exists, and then use ExcelWriter with mode='a' to append
    # If the file does not exist, mode='w' creates a new file
    if os.path.exists(excel_path):
        mode = 'a'
    else:
        mode = 'w'

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode) as writer:
        # Write the DataFrame to a specific sheet
        dict.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Created data dictionary successfully: saved to {excel_path}")

    return excel_path



def update_column_name(df, validate_dict_path):
    
    """
    Updates column name in the DataFrame.
    """
    
    # Load the necessary columns from the validate sheet
    validate_dict = pd.read_excel(validate_dict_path, sheet_name="validate", usecols=["Source Name", "Column Name"])

    # Create a dictionary from the DataFrame for renaming
    rename_dict = dict(zip(validate_dict['Source Name'], validate_dict['Column Name']))

    # Filter out keys that are not in the DataFrame's columns
    rename_dict = {k: v for k, v in rename_dict.items() if k in df.columns}

    # Rename columns in one go
    df = df.rename(columns=rename_dict)

    print("Updated column names successfully")

    return df



def add_date(df, time_column, date_column):
    df[time_column] = df[date_column].astype(str) + ' ' + df[time_column]
    return df



def replace_error(df, error_dir, file_name):
    
    # List error files
    error_files = sorted([f for f in os.listdir(error_dir) if f.endswith('.csv')])
    
    for i, error_file in enumerate(tqdm(error_files, desc="Processing error files"), start=1):
        # Extract column name from the file name
        column_name = error_file.replace(f'{file_name}_', '').replace('-error.csv', '')
        
        if column_name in df.columns:
            # Load the error CSV
            error_df = pd.read_csv(os.path.join(error_dir, error_file))
            
            # Create an array of 'error' where 'edit' is null
            null_values = set(error_df[error_df['Edit'].isna()]['Error'].values)
            mask = df[column_name].isin(null_values)
            df.loc[mask, column_name] = None
            
            # Create a mapping from error dictionary values
            error_dict = error_df.dropna(subset=['Edit']).set_index('Error')['Edit'].to_dict()
            
            # Replace values in the DataFrame
            df[column_name] = df[column_name].apply(lambda x: error_dict.get(x, x))
            
        else:
            print(f"Column {column_name} not found in DataFrame.")

    return df



def validate_input(input_value, data_type, format):
    """Validate the user's input against the expected data type and format, allowing explicit 'none' for null."""
    input_value = input_value.strip()
    if input_value.lower() in ['none', 'null', '']:
        return 'null'  # Special marker for intentional nulls

    try:
        if data_type == 'str' or data_type == 'object':
            return str(input_value)
        elif data_type in ['date', 'datetime', 'time']:
            return pd.to_datetime(input_value, format=format)
        elif data_type in ['int64', 'float64']:
            return pd.to_numeric(input_value)
    except ValueError:
        return None  # None still represents a conversion error



def update_column_type(df, dict_dir, file_name, output_dir, column_spc=None):
    
    # Import validate dict
    validate_dict_path = f'{dict_dir}/data_dict-{file_name}.xlsx'
    validate_dict = pd.read_excel(validate_dict_path, sheet_name="validate")
    
    # Create dataframe for manipulate
    df_mod = df.copy()
    i = 0
    e = 0
    total = len(validate_dict)
    
    if column_spc:
        validate_dict = validate_dict[validate_dict['Column Name'] == column_spc]

    error_dir = os.path.join(output_dir, f'error-{file_name}')
    if not os.path.exists(error_dir):
        os.makedirs(error_dir)
    else:
        error_files = sorted([f for f in os.listdir(error_dir) if f.endswith('.csv')])
        if len(error_files) > 0:
            print(f"Found {len(error_files)} files to process for error correction before updating column type.")
            df_mod = replace_error(df_mod, error_dir, file_name)

    for _, row in validate_dict.iterrows():
        i += 1
        column = row['Column Name']
        new_type = row['Data Type']
        format = row.get('Format', None)
        
        if column in df.columns:
        
            # Cleaning and preparing null values list
            if pd.notna(row['Null Value']):
                null_values = [x.replace('\xa0', ' ') for x in row['Null Value'].split(',')]
                null_values = [s.replace('"', '') for s in null_values]
            else:
                null_values = []  # Empty list if 'Null Value' is NaN
                
            # Replace specified null values in the DataFrame
            for null_value in null_values:
                print(f"Replace {null_value} in {column}")
                df_mod[column].replace({null_value: None}, inplace=True)  # Stripping spaces
            
            # For time-only data, aading the date to them
            if new_type == 'time':
                date_column = None
                while date_column == None:
                    date_column = input(f"Found the time data in {column}, please identify the date column for this column: ")
                    if date_column in df_mod.columns:
                        df_mod = add_date(df_mod, column, date_column)
                    else:
                        print(f"Column {date_column} not found in DataFrame.")
                        date_column = None
            
            errors = []
            
            try:
                if new_type in ['date', 'datetime', 'time']:
                    df_mod[column] = pd.to_datetime(df_mod[column], format=format)
                elif new_type in ['float64']:
                    df_mod[column] = pd.to_numeric(df_mod[column])
                elif new_type == 'str':
                    df_mod[column] = df_mod[column].astype('string')
                print(f"Successfully converted {column} to {new_type}")
            except:
                for idx, value in tqdm(df_mod[column].items(), total=df_mod[column].shape[0], desc=f"Failed converted {column} to {new_type} \nFinding the errors: "):
                    try:
                        if pd.isna(value):
                            continue  # Skip conversion for NaN values
                        if new_type == 'str' or new_type == 'object':
                            df_mod.at[idx, column] = str(value)
                        elif new_type in ['date', 'datetime', 'time']:
                            df_mod.at[idx, column] = pd.to_datetime(value, format=format)
                        elif new_type in ['int64', 'float64']:
                            df_mod.at[idx, column] = pd.to_numeric(value)
                    except (ValueError, TypeError):
                        errors.append({'Error': value})
    
        if len(errors) > 0:
            e += 1   
            error_df = pd.DataFrame(errors)
            
            error_path = os.path.join(error_dir, f"{e:02}_{file_name}_{column}-error.csv")
            error_df.to_csv(error_path, index=False)
            print(f"Found {len(error_df)} errors in {column}: saved to {error_path}")
                
    if e >= 1:
        print(f"There are errors in {e} columns and need to be modified")
        clean_path = f"Need error modification at {error_dir}"
    else:
        print(f"There is no errors in any columns. Saving the cleaned file --")           
        clean_path = f'{output_dir}/{file_name}_cleaned.csv'
        df_mod.to_csv(clean_path, index=False)
        print(f"Updated column data successfully without errors: processed data file were exported to {clean_path}")
        create_dictionary(df_mod, file_name, dict_dir)
    
    return df_mod, clean_path



def remove_duplicates(df, subset=None, keep='first'):
    """
    Remove duplicate rows from DataFrame.

    :param df: DataFrame from which duplicates will be removed.
    :param subset: Optional list of column names to consider for identifying duplicates.
    :param keep: Determines which duplicates (if any) to keep.
                 - 'first': (default) Drop duplicates except for the first occurrence.
                 - 'last': Drop duplicates except for the last occurrence.
                 - False: Drop all duplicates.
    :return: DataFrame with duplicates removed.
    """
    
    duplicates = df.duplicated(keep=keep)
    duplicate_count = duplicates.sum()
    
    print(f"Removing {duplicate_count} duplicates")
    
    df_unique = df.drop_duplicates(subset=subset, keep=keep)
    
    print(f"Removed {duplicate_count} duplicates successfully")
    
    return df_unique


def clean(file, data_dir, output_dir):
    df_raw, dict_path = data_import(file, data_dir)
    df = update_column_name(df_raw, dict_path)
    df, clean_path = update_column_type(df, data_dir, file, output_dir)
    df = remove_duplicates(df, keep='last')
    
    clear_output()
    
    # Cleaning summary
    clean_summary = {
        'File name': file,
        'Path to data dictionary': dict_path,
        'Path to cleaned dataset': clean_path
    }

    # Printing out the summary
    for key, value in clean_summary.items():
        print(f"{key:25}: {value}")
    
    return df, df_raw, dict_path


def str_split(input_str, pattern, number_of_groups):
    if isinstance(input_str, str):
        match = re.match(pattern, input_str)
        if match:
            return [match.group(i) for i in range(1, number_of_groups + 1)]
    return [None] * number_of_groups


def outlier_percentile(df, parameter, percentile=90, direction='more than'):
    # Calculate the specified percentile value
    values = df[parameter].dropna()
    if len(values) > 0:
        threshold = np.percentile(values, percentile)
    else:
        threshold = np.nan  # or handle it in another appropriate way
    
    # Set values to NaN based on the direction
    if direction == 'more than':
        df.loc[df[parameter] > threshold, parameter] = np.nan
    elif direction == 'less than':
        df.loc[df[parameter] < threshold, parameter] = np.nan
    elif direction == 'more than and equal to':
        df.loc[df[parameter] >= threshold, parameter] = np.nan
    elif direction == 'less than and equal to':
        df.loc[df[parameter] <= threshold, parameter] = np.nan
    else:
        raise ValueError("Direction must be 'more than', 'less than', 'more than and equal to', or 'less than and equal to'")
    
    return df





def mapping(df, parameter):
    map_dir = 'output/map'
    os.makedirs(map_dir, exist_ok=True)
    map_df_name = f'map_{parameter}'
    map_path = os.path.join(map_dir, f'{map_df_name}.csv')

    if os.path.exists(map_path):
        map_df = pd.read_csv(map_path)
        existing_mapped_values = set(map_df[parameter])
        unique_values = set(df[parameter].dropna().unique())
        
        new_values = unique_values - existing_mapped_values
        if new_values:
            print(f"New unique values found for {parameter}: {new_values}")
            if len(new_values) > 10:
                print(f"There are more than 10 new unique values for {parameter}. \nPlease fill in the new values in the CSV file and rerun the function.")
                new_map_df = pd.DataFrame({parameter: list(new_values), 'parameter_map': [None] * len(new_values)})
                new_map_df.to_csv(map_path, mode='a', header=False, index=False)
                return
            else:
                new_mappings = []
                for value in new_values:
                    input_value = input(f"Enter the map value for new unique {value} in parameter {parameter}: ")
                    new_mappings.append({parameter: value, 'parameter_map': input_value})
                
                # Convert new mappings to a DataFrame and concatenate
                new_map_df = pd.DataFrame(new_mappings)
                map_df = pd.concat([map_df, new_map_df], ignore_index=True)
                map_df.to_csv(map_path, index=False)
    else:
        print(f"Mapping file '{map_path}' not found. A new map will be created.")
        map_df = pd.DataFrame(columns=[parameter, 'parameter_map'])
        unique = df[parameter].dropna().unique()
        
        if len(unique) > 10:
            print(f"There are more than 10 unique values for the {parameter}. \nPlease fill in the values in the CSV file and rerun the function.")
            map_df = pd.concat([map_df, pd.DataFrame({parameter: unique, 'parameter_map': [None] * len(unique)})], ignore_index=True)
            map_df.to_csv(map_path, index=False)
            return
        else:
            mapping_dict = {}
            for i, uniq in enumerate(unique, start=1):
                input_value = input(f"Enter the map value for {uniq} in parameter {parameter} ({i}/{len(unique)}): ")
                if input_value.strip() == "":
                    input_value = None
                mapping_dict[uniq] = input_value

            map_df = pd.DataFrame(list(mapping_dict.items()), columns=[parameter, 'parameter_map'])
            map_df.to_csv(map_path, index=False)

    mapping_dict = map_df.set_index(parameter)['parameter_map'].to_dict()
    df[parameter] = df[parameter].map(mapping_dict)
    return df




    


def fill_missing(df, parameter, reference):
    map_dir = 'output/map'
    os.makedirs(map_dir, exist_ok=True)
    null = df[parameter].isna().sum()

    for ref in reference:
        map_df_name = f'map_{ref}-{parameter}'
        map_path = os.path.join(map_dir, f'{map_df_name}.csv')

        if os.path.exists(map_path):
            map_df = pd.read_csv(map_path)
        else:
            print(f"Mapping file '{map_path}' not found. A new map will be created.")
            map_df = pd.DataFrame(columns=[ref, parameter])

        mapping_dict = map_df.set_index(ref)[parameter].to_dict()

        df[parameter] = df[parameter].fillna(df[ref].map(mapping_dict))

        null_parameters = df[df[parameter].isna()][ref].unique()
        null_parameters = [param for param in null_parameters if pd.notna(param)]

        if len(null_parameters) > 10:
            print("There are more than 10 missing reference parameters. Please fill in the values in the CSV file.")
            null_df = pd.DataFrame(null_parameters, columns=[ref])
            null_df[parameter] = None
            
            if not map_df.empty:
                null_df = pd.concat([map_df, null_df]).drop_duplicates(subset=[ref]).reset_index(drop=True)
            
            null_df_path = os.path.join(map_dir, f'map_{ref}-{parameter}.csv')
            null_df.to_csv(null_df_path, index=False)
            print(f"Please fill in the missing values in {null_df_path} and rerun the function.")
            return
        
        elif len(null_parameters) > 0:
            for i, null_param in enumerate(null_parameters, start=1):
                input_value = input(f"Enter the value for {null_param} in parameter {ref} ({i}/{len(null_parameters)}): ")
                if input_value.strip() == "":
                    input_value = None

                # Update the mapping dictionary
                mapping_dict[null_param] = input_value

            # Re-apply the updated mapping to the DataFrame
            df[parameter] = df[parameter].fillna(df[ref].map(mapping_dict))
        
        
        # Export the updated mapping dictionary to the same map_path
        updated_map_df = pd.DataFrame(list(mapping_dict.items()), columns=[ref, parameter])
        updated_map_df.to_csv(map_path, index=False)

    null = df[parameter].isna().sum()
    percent_null = null / len(df) * 100
    print(f'Missing values in {parameter}: {null} ({percent_null:.2f}%)')

    return df

            